"""
reports.py
Professional and modular PDF/Excel reporting module for the laptop resale application.
Optimized for A5 Invoice layouts and proportional multi-page report headers.
"""

from __future__ import annotations

from io import BytesIO
from datetime import datetime
from typing import Any, Iterable, Mapping, Optional, Sequence

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4, A5, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.platypus import Frame, PageTemplate, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, NextPageTemplate
from openpyxl import Workbook
from models import get_all_settings
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.platypus import Macro

# Palette Warna Modern & Profesional
BIRU = colors.HexColor("#0d6efd")
BIRU_MUDA = colors.HexColor("#e7f1ff")
ABU_TUA = colors.HexColor("#212529")
ABU_GELAP = colors.HexColor("#495057")
BORDER = colors.HexColor("#dee2e6")

styles = getSampleStyleSheet()

# Tipografi Global
style_title = ParagraphStyle(
    "ReportTitle",
    parent=styles["Heading1"],
    fontName="Helvetica-Bold",
    fontSize=14,
    leading=18,
    alignment=TA_CENTER,
    textColor=BIRU,
    spaceAfter=4,
)
style_meta = ParagraphStyle(
    "MetaInfo",
    parent=styles["BodyText"],
    fontName="Helvetica",
    fontSize=8,
    leading=10,
    alignment=TA_CENTER,
    textColor=ABU_GELAP,
    spaceAfter=4,
)
style_normal = ParagraphStyle(
    "NormalBody",
    parent=styles["BodyText"],
    fontName="Helvetica",
    fontSize=8.5,
    leading=11,
    textColor=ABU_TUA,
)
style_normal_right = ParagraphStyle(
    "NormalBodyRight",
    parent=style_normal,
    alignment=TA_RIGHT,
)
style_th = ParagraphStyle(
    "TableHeader",
    parent=styles["BodyText"],
    fontName="Helvetica-Bold",
    fontSize=8.5,
    leading=11,
    textColor=colors.white,
)
style_th_right = ParagraphStyle(
    "TableHeaderRight",
    parent=style_th,
    alignment=TA_RIGHT,
)


def build_company_info(company_info: Optional[Mapping[str, Any]] = None) -> Mapping[str, str]:
    if company_info is None:
        company_info = get_all_settings()
    if isinstance(company_info, Mapping):
        return {
            "company_name": str(company_info.get("company_name") or "LaptopBekasApp"),
            "company_address": str(company_info.get("company_address") or "Alamat belum diisi"),
            "company_contact": str(company_info.get("company_contact") or "Kontak belum diisi"),
        }
    return {
        "company_name": "LaptopBekasApp",
        "company_address": "Alamat belum diisi",
        "company_contact": "Kontak belum diisi",
    }


def rp(value: Any) -> str:
    """Format nominal angka menjadi representasi mata uang Rupiah."""
    try:
        amount = float(value or 0)
    except (TypeError, ValueError):
        amount = 0
    return "Rp " + f"{amount:,.0f}".replace(",", ".")


class _PageCountingCanvas(pdf_canvas.Canvas):
    """Canvas internal untuk menghitung total halaman secara akurat."""
    page_count = 0

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        type(self).page_count = 0

    def showPage(self):
        type(self).page_count += 1
        super().showPage()


class BaseReport:
    """Base Engine PDF Report dengan Layout Frame Dinamis Multi-Halaman."""

    def __init__(
        self,
        company_info: Optional[Mapping[str, Any]] = None,
        report_title: str = "Laporan",
        period: str = "",
        data_source: str = "System",
        landscape_mode: bool = False,
        paper_size: tuple[float, float] = A4,
    ) -> None:
        self.company_info = build_company_info(company_info)
        self.report_title = report_title or "Laporan"
        self.period = period or "Semua data"
        self.data_source = data_source or "System"
        self.pagesize = landscape(paper_size) if landscape_mode else paper_size
        
        # Pengaturan Margin Luar Kertas
        self.left_margin = 10 * mm
        self.right_margin = 10 * mm
        self.top_margin = 10 * mm
        self.bottom_margin = 12 * mm
        
        self.total_pages = 1

    def build(self, story: Iterable[Any], buffer: Optional[BytesIO] = None) -> BytesIO:
        if buffer is None:
            buffer = BytesIO()

        printable_width = self.pagesize[0] - self.left_margin - self.right_margin
        
        first_page_frame = Frame(
            self.left_margin, 
            self.bottom_margin + 2 * mm, 
            printable_width, 
            self.pagesize[1] - self.bottom_margin - self.top_margin - 32 * mm,
            id="first_frame"
        )
        later_page_frame = Frame(
            self.left_margin, 
            self.bottom_margin + 2 * mm, 
            printable_width, 
            self.pagesize[1] - self.bottom_margin - self.top_margin - 8 * mm,
            id="later_frame"
        )

        _PageCountingCanvas.page_count = 0

        doc = SimpleDocTemplate(
            buffer,
            pagesize=self.pagesize,
            leftMargin=self.left_margin,
            rightMargin=self.right_margin,
            topMargin=self.top_margin,
            bottomMargin=self.bottom_margin,
            canvasmaker=_PageCountingCanvas,
        )

        doc.addPageTemplates([
            PageTemplate(id="FirstPage", frames=[first_page_frame], onPage=self._draw_first_page),
            PageTemplate(id="LaterPages", frames=[later_page_frame], onPage=self._draw_later_page),
        ])

        story_list = [NextPageTemplate("LaterPages")]
        story_list.extend(list(story))

        doc.build(story_list)
        self.total_pages = max(1, _PageCountingCanvas.page_count)
        buffer.seek(0)
        return buffer

    def _draw_first_page(self, canvas: pdf_canvas.Canvas, doc: Any) -> None:
        canvas.saveState()
        # Menggambar Header Utama hanya di Halaman Pertama
        y = self.pagesize[1] - self.top_margin
        
        canvas.setFont("Helvetica-Bold", 12)
        canvas.setFillColor(BIRU)
        canvas.drawString(self.left_margin, y - 2 * mm, self.company_info["company_name"])
        
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(ABU_GELAP)
        canvas.drawString(self.left_margin, y - 6 * mm, self.company_info["company_address"])
        canvas.drawString(self.left_margin, y - 10 * mm, self.company_info["company_contact"])

        canvas.setFont("Helvetica-Bold", 12)
        canvas.setFillColor(ABU_TUA)
        canvas.drawCentredString(self.pagesize[0] / 2, y - 16 * mm, self.report_title)

        canvas.setFont("Helvetica", 8)
        canvas.drawCentredString(self.pagesize[0] / 2, y - 21 * mm, f"Periode: {self.period}  |  Sumber: {self.data_source}")
        
        canvas.setStrokeColor(BORDER)
        canvas.setLineWidth(0.5)
        canvas.line(self.left_margin, y - 25 * mm, self.pagesize[0] - self.right_margin, y - 25 * mm)
        
        canvas.restoreState()
        self.draw_footer(canvas, doc)

    def _draw_later_page(self, canvas: pdf_canvas.Canvas, doc: Any) -> None:
        canvas.saveState()
        # Header minimalis untuk halaman 2 dan seterusnya
        y = self.pagesize[1] - self.top_margin
        canvas.setFont("Helvetica-Bold", 8)
        canvas.setFillColor(ABU_GELAP)
        canvas.drawString(self.left_margin, y, self.report_title)
        canvas.drawRightString(self.pagesize[0] - self.right_margin, y, f"Periode: {self.period}")
        canvas.setStrokeColor(BORDER)
        canvas.setLineWidth(0.4)
        canvas.line(self.left_margin, y - 2 * mm, self.pagesize[0] - self.right_margin, y - 2 * mm)
        canvas.restoreState()
        self.draw_footer(canvas, doc)

    def draw_footer(self, canvas: pdf_canvas.Canvas, doc: Any) -> None:
        canvas.saveState()
        y = self.bottom_margin - 4 * mm
        canvas.setStrokeColor(BORDER)
        canvas.setLineWidth(0.4)
        canvas.line(self.left_margin, y + 4 * mm, self.pagesize[0] - self.right_margin, y + 4 * mm)
        
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(ABU_GELAP)
        canvas.drawString(self.left_margin, y, f"Dicetak: {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}")
        canvas.drawCentredString(self.pagesize[0] / 2, y, "Laporan Sistem LaptopBekasApp")
        canvas.drawRightString(self.pagesize[0] - self.right_margin, y, f"Halaman {canvas.getPageNumber()} dari {self.total_pages}")
        canvas.restoreState()

    def build_table(
        self,
        headers: Sequence[str],
        rows: Sequence[Sequence[Any]],
        numeric_columns: Optional[Sequence[int]] = None,
        col_ratios: Optional[Sequence[float]] = None,
    ) -> Table:
        """Membangun tabel data yang terbungkus otomatis (auto-wrapped text)."""
        num_cols = len(headers)
        num_cols = max(num_cols, 1)
        numeric_cols = numeric_columns or []

        # Proses data header menjadi Paragraph agar tidak overflow
        header_row = []
        for idx, h in enumerate(headers):
            style = style_th_right if idx in numeric_cols else style_th
            header_row.append(Paragraph(str(h), style))

        data = [header_row]

        # Proses row data menjadi Paragraph
        for r in rows:
            formatted_row = []
            for idx, cell in enumerate(r):
                style = style_normal_right if idx in numeric_cols else style_normal
                formatted_row.append(Paragraph(str(cell), style))
            data.append(formatted_row)

        # Menghitung proporsi lebar kolom yang adil
        printable_width = self.pagesize[0] - self.left_margin - self.right_margin
        if col_ratios and len(col_ratios) == num_cols:
            total_ratio = sum(col_ratios)
            col_widths = [(r / total_ratio) * printable_width for r in col_ratios]
        else:
            col_widths = [printable_width / num_cols] * num_cols

        table = Table(data, repeatRows=1, colWidths=col_widths, hAlign="LEFT")
        
        t_style = [
            ("BACKGROUND", (0, 0), (-1, 0), BIRU),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, BIRU_MUDA]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]
        table.setStyle(TableStyle(t_style))
        return table


def buat_nota_transaksi_a5(
    jenis: str, 
    transaksi: Mapping[str, Any], 
    items: Sequence[Mapping[str, Any]], 
    company_info: Optional[Mapping[str, Any]] = None
) -> BytesIO:
    """
    Fungsi Master Pembuatan Nota Transaksi Penjualan / Pembelian berukuran A5 (Kertas Kasir Umum).
    Mendukung penuh pemisahan block data Header Metadata dan Detail Item.
    """
    judul_nota = "NOTA PENJUALAN" if jenis.upper() == "PENJUALAN" else "NOTA PEMBELIAN"
    sumber_data = "Data Penjualan" if jenis.upper() == "PENJUALAN" else "Data Pembelian"
    
    report = BaseReport(
        company_info=company_info or {
            "company_name": "LaptopBekasApp",
            "company_address": "Pusat Perdagangan Laptop Seken, Indonesia",
            "company_contact": "Telp/WA: 0812-3456-7890",
        },
        report_title=judul_nota,
        period=str(transaksi.get("tanggal") or "-"),
        data_source=sumber_data,
        paper_size=A5,
        landscape_mode=True  # Landscape A5 sangat ideal untuk layout invoice horizontal
    )

    story = []

    # 1. BLOK HEADER NOTA (Pihak Terkait & Metadata Administrasi)
    entitas_label = "Pelanggan / Pembeli :" if jenis.upper() == "PENJUALAN" else "Supplier / Vendor :"
    entitas_nama = transaksi.get("pembeli") if jenis.upper() == "PENJUALAN" else transaksi.get("supplier")
    
    meta_data = [
        [Paragraph(f"<b>{entitas_label}</b>", style_normal), Paragraph(f"<b>Nomor Transaksi:</b>", style_normal)],
        [Paragraph(str(entitas_nama or "-"), style_normal), Paragraph(str(transaksi.get("nomor_penjualan") or transaksi.get("nomor_pembelian") or "-"), style_normal)],
        [Paragraph(f"Metode Pembayaran: {transaksi.get('metode_pembayaran') or '-'}", style_normal), Paragraph(f"Tanggal: {transaksi.get('tanggal') or '-'}", style_normal)]
    ]
    
    printable_width = report.pagesize[0] - report.left_margin - report.right_margin
    meta_table = Table(meta_data, colWidths=[printable_width * 0.5, printable_width * 0.5])
    meta_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
    ]))
    
    story.append(meta_table)
    story.append(Spacer(1, 4 * mm))

    # 2. BLOK DETAIL ITEMS (Tabel Item List Barang)
    headers = ["No", "Deskripsi Barang / Seri Laptop", "Kondisi", "Qty", "Harga Satuan", "Total"]
    rows = []
    
    for idx, item in enumerate(items, start=1):
        harga_satuan = float(item.get("harga_jual") if jenis.upper() == "PENJUALAN" else item.get("harga_beli") or 0)
        qty = int(item.get("qty") or 1)
        subtotal = harga_satuan * qty
        
        rows.append([
            str(idx),
            f"{item.get('merk', '')} {item.get('seri', '')}".strip() or "Laptop Item",
            str(item.get("kondisi") or "-"),
            str(qty),
            rp(harga_satuan),
            rp(subtotal)
        ])

    # Proporsi lebar kolom (No kecil, deskripsi barang besar)
    ratios = [0.8, 4.0, 1.5, 0.8, 2.0, 2.0]
    story.append(report.build_table(headers, rows, numeric_columns=[3, 4, 5], col_ratios=ratios))
    story.append(Spacer(1, 3 * mm))

    # 3. BLOK SUMMARY TOTAL (Bawah Kanan)
    total_akhir = float(transaksi.get("harga_jual") or transaksi.get("total_modal") or 0)
    
    summary_rows = []
    if jenis.upper() == "PENJUALAN" and "laba" in transaksi:
        summary_rows.append([Paragraph("Total Nilai Transaksi:", style_normal), Paragraph(rp(total_akhir), style_normal_right)])
    else:
        summary_rows.append([Paragraph("Total Modal Pembelian:", style_normal), Paragraph(rp(total_akhir), style_normal_right)])

    summary_table = Table(summary_rows, colWidths=[printable_width * 0.65, printable_width * 0.35])
    summary_table.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "RIGHT"),
        ("LINEABOVE", (0, 0), (-1, 0), 0.8, BIRU),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(summary_table)
    
    if transaksi.get("catatan"):
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(f"<b>Catatan:</b> {transaksi.get('catatan')}", style_normal))

    return report.build(story)


def buat_nota_pembelian_pdf(laptop: Any) -> BytesIO:
    """Wrapper backwards compatibility nota pembelian berbentuk objek tunggal."""
    transaksi = {
        "tanggal": _get_value(laptop, "tanggal"),
        "nomor_pembelian": _get_value(laptop, "nomor_pembelian"),
        "supplier": _get_value(laptop, "supplier"),
        "metode_pembayaran": "Tunai/Transfer",
        "total_modal": _get_value(laptop, "total_modal"),
        "catatan": _get_value(laptop, "catatan")
    }
    items = [{
        "merk": _get_value(laptop, "merk"),
        "seri": _get_value(laptop, "seri"),
        "kondisi": _get_value(laptop, "kondisi"),
        "qty": 1,
        "harga_beli": _get_value(laptop, "harga_beli")
    }]
    return buat_nota_transaksi_a5("PEMBELIAN", transaksi, items)


def buat_nota_penjualan_pdf(penjualan: Any, laptop: Any) -> BytesIO:
    """Wrapper backwards compatibility nota penjualan berbentuk objek tunggal."""
    transaksi = {
        "tanggal": _get_value(penjualan, "tanggal"),
        "nomor_penjualan": _get_value(penjualan, "nomor_penjualan"),
        "pembeli": _get_value(penjualan, "pembeli"),
        "metode_pembayaran": _get_value(penjualan, "metode_pembayaran"),
        "harga_jual": _get_value(penjualan, "harga_jual"),
        "laba": _get_value(penjualan, "laba"),
        "catatan": _get_value(penjualan, "catatan")
    }
    items = [{
        "merk": _get_value(laptop, "merk"),
        "seri": _get_value(laptop, "seri"),
        "kondisi": _get_value(laptop, "kondisi"),
        "qty": 1,
        "harga_jual": _get_value(penjualan, "harga_jual")
    }]
    return buat_nota_transaksi_a5("PENJUALAN", transaksi, items)


def buat_laporan_pdf(
    judul: str,
    subjudul: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    total_row: Optional[Sequence[Any]] = None,
    company_info: Optional[Mapping[str, Any]] = None,
    report_title: Optional[str] = None,
    period: Optional[str] = None,
    data_source: Optional[str] = None,
) -> BytesIO:
    """Membuat PDF laporan berskala besar (A4) yang rapih di setiap halamannya."""
    report = BaseReport(
        company_info=company_info,
        report_title=report_title or judul,
        period=period or subjudul or "Semua data",
        data_source=data_source or "System",
        paper_size=A4
    )
    story = [Spacer(1, 2 * mm)]
    
    all_rows = list(rows)
    if total_row:
        all_rows.append(total_row)
        
    numeric_columns = [len(headers) - 1] if headers else []
    
    # Memastikan build_table mengembalikan objek Table (bukan fungsi)
    story.append(report.build_table(headers, all_rows, numeric_columns=numeric_columns))
    
    return report.build(story) # Baris 474

def buat_laporan_excel(
    judul: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    total_row: Optional[Sequence[Any]] = None,
    company_info: Optional[Mapping[str, Any]] = None,
    period: Optional[str] = None,
    data_source: Optional[str] = None,
) -> BytesIO:
    """Membuat file spreadsheet Excel dengan auto-column width & alignment numerik yang rapih."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan"

    max_col = max(len(headers), 1)
    
    # Header Judul Excel
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell = ws.cell(row=1, column=1, value=judul)
    cell.font = Font(name="Arial", bold=True, size=14, color="0D6EFD")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25

    # Sub-header info metadata tersemat
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    sub_text = f"Periode: {period or 'Semua data'} | Sumber: {data_source or 'System'}"
    sub_cell = ws.cell(row=2, column=1, value=sub_text)
    sub_cell.font = Font(name="Arial", size=9, italic=True, color="495057")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Table Header Styling
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    
    for idx, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[4].height = 22

    # Menulis Data Baris
    current_row = 5
    for r in rows:
        for idx, val in enumerate(r, start=1):
            c = ws.cell(row=current_row, column=idx, value=val)
            c.font = Font(name="Arial", size=10)
            # Jika isinya berupa angka, geser rata kanan
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        current_row += 1

    # Menulis Data Baris Total Kumulatif
    if total_row:
        for idx, val in enumerate(total_row, start=1):
            c = ws.cell(row=current_row, column=idx, value=val)
            c.font = Font(name="Arial", bold=True, size=10)
            c.fill = PatternFill(start_color="E7F1FF", end_color="E7F1FF", fill_type="solid")
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right", vertical="center")
            else:
                c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[current_row].height = 20

    # Auto-adjust lebar kolom secara proporsional agar tidak ada teks terpotong
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row > 2 and cell.value:  
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _get_value(data: Any, key: str, default: Any = None) -> Any:
    if data is None:
        return default
    if isinstance(data, Mapping):
        return data.get(key, default)
    if hasattr(data, key):
        return getattr(data, key)
    if hasattr(data, "__getitem__"):
        try:
            return data[key]
        except Exception:
            pass
    return default